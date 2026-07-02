from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import AuctionItem


def save_items_to_excel(items: Iterable[AuctionItem], output_path: str | Path) -> Path:
    rows = [item.normalized() for item in items]
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "경매물건"

    headers = _collect_headers(rows)
    if not headers:
        headers = ["수집결과"]
        rows = [{"수집결과": "수집된 데이터가 없습니다."}]

    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])

    _style_sheet(sheet, headers)
    workbook.save(output)
    return output


def _collect_headers(rows: list[dict[str, str]]) -> list[str]:
    preferred = [
        "사건번호",
        "물건번호",
        "소재지",
        "용도",
        "감정평가액",
        "최저매각가격",
        "매각기일",
        "진행상태",
    ]
    seen: set[str] = set()
    headers: list[str] = []

    for header in preferred:
        if any(header in row for row in rows):
            headers.append(header)
            seen.add(header)

    for row in rows:
        for header in row:
            if header not in seen:
                headers.append(header)
                seen.add(header)
    return headers


def _style_sheet(sheet, headers: list[str]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    border_alignment = Alignment(vertical="top", wrap_text=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = border_alignment

    for index, header in enumerate(headers, start=1):
        letter = get_column_letter(index)
        max_len = len(header)
        for cell in sheet[letter]:
            max_len = max(max_len, len(str(cell.value or "")))
        sheet.column_dimensions[letter].width = min(max(max_len + 2, 12), 45)

